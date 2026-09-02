"""
invoice_radar.py
────────────────
Invoice Radar / Missing Invoices report.

Source: notebook prototype promoted to Composer (sanitized)

Queries LPV vs invoice discrepancies, writes ``bi.all_invoices``,
builds a 4-sheet Excel file, and returns an email payload for the shared
``email_delivery`` module.

Environment variables (set by Airflow before generate_reports() runs):
  GCP_PROJECT_ID, BQ_SOURCE_VIEW, BQ_ASSET_TABLE, BQ_PRICING_TABLE,
  BQ_DEST_FULL, WRITE_ALL_INVOICES,
  EMAIL_FROM, EMAIL_FROM_NAME, EMAIL_RECIPIENTS, EMAIL_CC (optional)
  EMAIL_PROVIDER — "smtp" (default) or "sendgrid" (used by main() local runs)
  SMTP_* / SENDGRID_API_KEY — see email_delivery module
"""

from __future__ import annotations

import calendar
import io
import logging
import os
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from dotenv import load_dotenv

    load_dotenv(Path(__file__).parent / ".env")
except ImportError:
    pass

# ── Logging ───────────────────────────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

TEMPLATE_DIR = Path(__file__).parent
TEMPLATE_FILE = "invoice_radar_alert.html"
BUSINESS_TZ = ZoneInfo("Europe/Amsterdam")

# ── Config ────────────────────────────────────────────────────────────────────


def _env(key: str, required: bool = True, default: str = "") -> str:
    value = os.getenv(key, default)
    if required and not value:
        raise EnvironmentError(f"Required environment variable '{key}' is not set.")
    return value


def _tables() -> dict[str, str]:
    """Fully-qualified BigQuery references from env (set by Airflow)."""
    return {
        "source_view": _env("BQ_SOURCE_VIEW"),
        "asset_table": _env("BQ_ASSET_TABLE"),
        "pricing_table": _env("BQ_PRICING_TABLE"),
        "dest_full": _env("BQ_DEST_FULL"),
        "project": _env("GCP_PROJECT_ID"),
    }


def _write_all_invoices_enabled() -> bool:
    return _env("WRITE_ALL_INVOICES", required=False, default="true").lower() in (
        "1",
        "true",
        "yes",
    )


# ── Date helpers ──────────────────────────────────────────────────────────────


def get_report_dates() -> dict[str, str]:
    # Composer workers are UTC; use Amsterdam so the D-3 window matches the DAG.
    today = datetime.now(BUSINESS_TZ).date()
    report_date = today - timedelta(days=3)
    next_day = today - timedelta(days=2)
    next_next_day = today - timedelta(days=1)

    y, m, d_ = report_date.year, report_date.month, report_date.day
    if m == 12:
        next_month_report = date(y + 1, 1, d_)
    else:
        last_day = calendar.monthrange(y, m + 1)[1]
        next_month_report = date(y, m + 1, min(d_, last_day))

    cal_month_start = report_date.replace(day=1)
    cal_month_end = next_next_day.replace(day=1)

    if cal_month_start == cal_month_end:
        cal_month_filter = f"AND cal_month = '{cal_month_start.isoformat()}'"
    else:
        cal_month_filter = (
            f"AND cal_month IN "
            f"('{cal_month_start.isoformat()}', '{cal_month_end.isoformat()}')"
        )

    return {
        "report_date": report_date.isoformat(),
        "next_day": next_day.isoformat(),
        "next_next_day": next_next_day.isoformat(),
        "next_month_report": next_month_report.isoformat(),
        "cal_month_filter": cal_month_filter,
        "report_date_label": report_date.strftime("%d %B %Y"),
    }


def fmt_eur(value: float | None) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "—"
    return f"{value:,.2f} €".replace(",", "X").replace(".", ",").replace("X", ".")


# ── BigQuery client ───────────────────────────────────────────────────────────


def get_bq_client(project: str) -> bigquery.Client:
    sa = os.environ.get("GCP_SERVICE_ACCOUNT")
    if sa:
        from google.oauth2 import service_account

        creds = service_account.Credentials.from_service_account_file(sa)
        return bigquery.Client(project=project, credentials=creds)
    return bigquery.Client(project=project)


def run_query(client: bigquery.Client, sql: str, project: str) -> pd.DataFrame:
    """Run SQL via the REST row iterator (avoids BigQuery Storage API)."""
    job_config = bigquery.QueryJobConfig(use_legacy_sql=False)
    job = client.query(sql, job_config=job_config, project=project)
    rows = list(job.result())
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame([dict(row.items()) for row in rows])
    for col in ("inv_revenue", "lpv_revenue", "abs_diff_eur"):
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


# ── Source SQL — mirrors bigquery.js / GitLab invoice_radar.py exactly ────────


def build_source_sql(d: dict[str, str], tables: dict[str, str]) -> str:
    source_view = tables["source_view"]
    asset_table = tables["asset_table"]
    pricing_table = tables["pricing_table"]
    return f"""
WITH invoices AS (
  SELECT
    l.cal_month, l.country_code, l.establishment_uid, l.sales_order, l.sol_id,
    l.partner_id, l.product_code, l.product_is_setup, l.commitment_period,
    l.billing_period, l.recurring_revenue, l.activation_revenue,
    l.inv_activation_rev, l.inv_recurring_rev, l.lpv_revenue, l.inv_revenue,
    l.nr_invoices, l.diff_eur, l.diff_activation_eur, l.diff_recurring_eur,
    l.abs_diff_eur, l.diff_pct_vs_lpv, l.diff_pct_vs_invoice,
    l.reconciliation_status, l.discrepancy_reason, l.discrepancy_risk,
    l.match_est_prod_lev, l.vat_info, l.has_vat,
    l.start_date, l.end_date, l.start_suspension_date, l.next_invoice_date,
    l.last_invoice_date, l.months_to_next_invoice, l.first_onb_date_ops,
    l.onb_month_ops, l.initial_order_date, l.subscription_management_new,
    l.effective_pricing_date,
    -- prod lpv_adjusted has no snap_start/snap_end; keep aliases for reason logic
    CAST(NULL AS DATE) AS snap_start,
    CAST(NULL AS DATE) AS snap_end,
    a.asset_activated_by,
    a.asset_activated_by_email,
    CASE
      WHEN LOWER(l.product_code) LIKE '%pos%'
        OR l.product_code IN ('PI_12IP_GT8000','PI_12IP_GT6000','PI_12IP_GT4000',
                              'PI_12IP_GT2000','PI_12IP_LT2000','PI_4IP_GT4000',
                              'PI_4IP_GT2000','PayInst')
      THEN 'POS'
      WHEN LOWER(l.product_code) LIKE 'pt_%'
        OR LOWER(l.product_code) LIKE 'fo_b2b2c_tf%'
      THEN 'PAYMENTS'
      WHEN LOWER(l.product_code) LIKE '%mto_%'
        OR LOWER(l.product_code) LIKE 'wsre%'
        OR LOWER(l.product_code) LIKE 'ubi_%'
        OR LOWER(l.product_code) LIKE '%prm%'
        OR l.product_code LIKE 'FO_FOBAS%'
        OR l.product_code LIKE '%LEGACY_PACK%'
      THEN 'LEGACY_SUITE'
      ELSE 'Other'
    END AS main_category
  FROM `{source_view}` l
  LEFT JOIN (
    SELECT DISTINCT so_name, asset_activated_by, asset_activated_by_email
    FROM `{asset_table}`
    QUALIFY ROW_NUMBER() OVER (PARTITION BY so_name ORDER BY asset_created_date DESC) = 1
  ) a ON l.sales_order = a.so_name
  WHERE (l.lpv_revenue > 0 OR l.inv_revenue > 0)
    AND l.start_date <= CURRENT_DATE()
    AND (
      l.next_invoice_date = '{d["report_date"]}'
      OR l.last_invoice_date = '{d["report_date"]}'
      OR (l.last_invoice_date = '{d["next_day"]}'
          AND l.next_invoice_date = DATE_ADD('{d["report_date"]}', INTERVAL 1 MONTH))
      OR (l.last_invoice_date = '{d["next_next_day"]}'
          AND l.next_invoice_date = DATE_ADD('{d["report_date"]}', INTERVAL 1 MONTH))
    )
    {d["cal_month_filter"]}
    AND (l.end_date IS NULL OR l.end_date > CURRENT_DATE())
    AND (l.start_suspension_date IS NULL OR l.start_suspension_date > CURRENT_DATE())
),

pricing AS (
  SELECT
    ProductCode_Lookup_Odoo AS product_code,
    country,
    Commitment_Period        AS commitment_period,
    Billing_Period           AS billing_period,
    Type                     AS product_type,
    list_price_reg,
    price_start_date,
    price_end_date
  FROM `{pricing_table}`
  QUALIFY ROW_NUMBER() OVER (
    PARTITION BY ProductCode_Lookup_Odoo, country, Commitment_Period,
                 Billing_Period, price_start_date, price_end_date, Type
    ORDER BY list_price_reg DESC
  ) = 1
),

invoice_pricing AS (
  SELECT
    i.establishment_uid, i.sales_order, i.sol_id,
    i.product_code, i.cal_month, i.product_is_setup,
    p.list_price_reg   AS p_lpv_price,
    p.price_start_date AS p_lpv_start,
    p.price_end_date   AS p_lpv_end,
    p.product_code     AS p_lpv_code
  FROM invoices i
  LEFT JOIN pricing p
    ON  i.product_code      = p.product_code
    AND i.country_code      = p.country
    AND i.commitment_period = p.commitment_period
    AND i.billing_period    = p.billing_period
    AND (
      (i.product_is_setup = TRUE  AND p.product_type = 'Activation')
      OR (i.product_is_setup = FALSE AND p.product_type = 'Recurring')
      OR i.product_is_setup IS NULL
    )
    AND COALESCE(i.effective_pricing_date, i.start_date) >= p.price_start_date
    AND COALESCE(i.effective_pricing_date, i.start_date) <= COALESCE(p.price_end_date, DATE '2999-12-31')
  QUALIFY ROW_NUMBER() OVER (
    PARTITION BY i.establishment_uid, i.sales_order, i.sol_id,
                 i.product_code, i.cal_month, i.product_is_setup
    ORDER BY p.price_start_date ASC
  ) = 1
),

pricing_any AS (
  SELECT
    product_code, country, commitment_period, billing_period,
    product_type, list_price_reg, price_start_date, price_end_date
  FROM pricing
  QUALIFY ROW_NUMBER() OVER (
    PARTITION BY product_code, country, commitment_period, billing_period, product_type
    ORDER BY price_start_date DESC
  ) = 1
)

SELECT
  i.*,
  CASE
    -- 1. Sentinel onboarding date
    WHEN i.lpv_revenue = 0 AND i.inv_revenue > 0
      AND i.first_onb_date_ops = '2016-01-01'
    THEN CONCAT(
      'Invoiced but not onboarded: onboarding date is sentinel (2016-01-01)',
      ' — LPV cannot calculate revenue without a valid onboarding date.'
    )
    -- 2. Renewal not yet started
    WHEN i.lpv_revenue = 0 AND i.inv_revenue > 0
      AND i.start_date > DATE_TRUNC(i.cal_month, MONTH)
      AND i.start_date > CURRENT_DATE()
    THEN CONCAT(
      'Renewal not yet started: start_date ', CAST(i.start_date AS STRING),
      ' is after today ', CAST(CURRENT_DATE() AS STRING),
      ' — LPV snap not yet active for this month'
    )
    -- Contract ended but invoice still raised
    WHEN i.lpv_revenue = 0 AND i.inv_revenue > 0
      AND i.end_date < i.cal_month
    THEN CONCAT(
      'Contract ended: end_date ', CAST(i.end_date AS STRING),
      ' is before billing month ', CAST(i.cal_month AS STRING),
      ' — LPV excludes closed subscriptions but invoice was still raised'
    )
    -- 3a. Paused snap covers month but asset resumed too late for LPV to capture
    WHEN i.lpv_revenue = 0 AND i.inv_revenue > 0
      AND i.start_suspension_date IS NOT NULL
      AND i.start_date > DATE_TRUNC(i.cal_month, MONTH)
    THEN CONCAT(
      'Asset paused during billing month and resumed after month started: ',
      'start_date ', CAST(i.start_date AS STRING),
      ' — LPV snap active from next month only'
    )
    -- 3b. Asset paused
    WHEN i.lpv_revenue = 0 AND i.inv_revenue > 0
      AND i.start_suspension_date IS NOT NULL
    THEN 'Asset paused: subscription suspended for this billing period'
    -- 4. No pricing entry exists at all
    WHEN i.lpv_revenue = 0 AND i.inv_revenue > 0
      AND pa.product_code IS NULL
    THEN CONCAT(
      'No pricing entry found: ',
      i.product_code, ' / ', i.country_code,
      ' / commitment=', CAST(i.commitment_period AS STRING),
      ' / billing=', CAST(i.billing_period AS STRING)
    )
    -- 5. Pricing exists but effective_pricing_date outside every known window
    WHEN i.lpv_revenue = 0 AND i.inv_revenue > 0
      AND pa.product_code IS NOT NULL
      AND ip.p_lpv_code IS NULL
      AND i.initial_order_date IS NOT NULL
    THEN CONCAT(
      'Pricing expired or not yet valid: ',
      i.product_code,
      ' - effective_pricing_date ', CAST(COALESCE(i.effective_pricing_date, i.start_date) AS STRING),
      ' outside latest price window ',
      CAST(pa.price_start_date AS STRING),
      ' → ', COALESCE(CAST(pa.price_end_date AS STRING), 'open')
    )
    -- 5b. SOL_ID mismatch
    WHEN i.lpv_revenue = 0 AND i.inv_revenue > 0
      AND i.first_onb_date_ops IS NULL
      AND i.initial_order_date IS NULL
    THEN CONCAT(
      'SOL_ID mismatch: invoice sol_id ', CAST(i.sol_id AS STRING),
      ' has no matching record in LPV ops for sales_order ', i.sales_order
    )
    -- 5c. SnapStart in future
    WHEN i.lpv_revenue = 0 AND i.inv_revenue > 0
      AND i.snap_start IS NOT NULL
      AND i.snap_start > DATE_TRUNC(i.cal_month, MONTH)
    THEN CONCAT(
      'LPV snap not yet active: subscription has SnapStart ',
      CAST(i.snap_start AS STRING),
      ' which is after billing month ', CAST(i.cal_month AS STRING),
      '. Likely a change_of_plan or renewal — LPV will capture from ',
      FORMAT_DATE('%B %Y', DATE_TRUNC(i.snap_start, MONTH))
    )
    -- 5d. SnapEnd before cal_month — subscription closed
    WHEN i.lpv_revenue = 0 AND i.inv_revenue > 0
      AND i.snap_end IS NOT NULL
      AND i.snap_end < DATE_TRUNC(i.cal_month, MONTH)
    THEN CONCAT(
      'Contract closed in LPV ops: all snap windows ended before billing month ',
      CAST(i.cal_month AS STRING),
      '. Last snap_end: ', CAST(i.snap_end AS STRING),
      ', invoice raised after contract closure.'
    )
    -- 6. catch-all lpv=0
    WHEN i.lpv_revenue = 0 AND i.inv_revenue > 0
    THEN CASE
      WHEN i.first_onb_date_ops IS NULL
        AND i.initial_order_date IS NULL
        AND i.start_date < DATE_TRUNC(i.cal_month, MONTH)
      THEN CONCAT(
        'LPV revenue is 0: subscription likely closed or expired before billing month.',
        ' All LPV ops snap windows ended before cal_month ', CAST(i.cal_month AS STRING),
        ' — invoice raised after contract closure.',
        ' start_date: ', CAST(i.start_date AS STRING),
        ', end_date (Odoo): ', CAST(i.end_date AS STRING)
      )
      WHEN i.first_onb_date_ops IS NULL
        AND i.initial_order_date IS NULL
        AND i.start_date >= DATE_TRUNC(i.cal_month, MONTH)
      THEN CONCAT(
        'LPV revenue is 0: new subscription not yet captured in LPV snapshot.',
        ' Possible SnapEnd sentinel (2016-01-01) or SOL_ID mismatch.',
        ' start_date: ', CAST(i.start_date AS STRING),
        ', cal_month: ', CAST(i.cal_month AS STRING)
      )
      ELSE CONCAT(
        'LPV revenue is 0: subscription snap window not active for this billing month.',
        ' Possible causes: change_of_plan/renewal with future SnapStart,',
        ' or asset paused at LPV stage level.',
        ' start_date: ', CAST(i.start_date AS STRING),
        ', first_onb_date: ', CAST(i.first_onb_date_ops AS STRING)
      )
    END
    -- 7. Under-invoiced
    WHEN i.lpv_revenue > 0 AND i.inv_revenue > 0 AND i.inv_revenue < i.lpv_revenue
    THEN CONCAT(
      'Under-invoiced: LPV=', FORMAT('%.2f', i.lpv_revenue), '€',
      ' vs Invoice=', FORMAT('%.2f', i.inv_revenue), '€'
    )
    -- 8. Over-invoiced
    WHEN i.lpv_revenue > 0 AND i.inv_revenue > i.lpv_revenue
    THEN CONCAT(
      'Over-invoiced: LPV=', FORMAT('%.2f', i.lpv_revenue), '€',
      ' vs Invoice=', FORMAT('%.2f', i.inv_revenue), '€'
    )
    -- 9. Fallback
    ELSE i.discrepancy_reason
  END AS enriched_discrepancy_reason

FROM invoices i
LEFT JOIN invoice_pricing ip
  ON  i.establishment_uid = ip.establishment_uid
  AND i.sales_order       = ip.sales_order
  AND i.sol_id            = ip.sol_id
  AND i.product_code      = ip.product_code
  AND i.cal_month         = ip.cal_month
  AND i.product_is_setup  = ip.product_is_setup
LEFT JOIN pricing_any pa
  ON  i.product_code      = pa.product_code
  AND i.country_code      = pa.country
  AND i.commitment_period = pa.commitment_period
  AND i.billing_period    = pa.billing_period
  AND (
    (i.product_is_setup = TRUE  AND pa.product_type = 'Activation')
    OR (i.product_is_setup = FALSE AND pa.product_type = 'Recurring')
    OR i.product_is_setup IS NULL
  )
"""


# ── Write all_invoices to BigQuery (WRITE_TRUNCATE) ───────────────────────────

ALL_INVOICES_COLUMNS = [
    "cal_month",
    "country_code",
    "establishment_uid",
    "sales_order",
    "sol_id",
    "partner_id",
    "product_code",
    "product_is_setup",
    "commitment_period",
    "billing_period",
    "recurring_revenue",
    "activation_revenue",
    "inv_activation_rev",
    "inv_recurring_rev",
    "lpv_revenue",
    "inv_revenue",
    "nr_invoices",
    "diff_eur",
    "diff_activation_eur",
    "diff_recurring_eur",
    "abs_diff_eur",
    "diff_pct_vs_lpv",
    "diff_pct_vs_invoice",
    "reconciliation_status",
    "discrepancy_reason",
    "discrepancy_risk",
    "match_est_prod_lev",
    "vat_info",
    "has_vat",
    "start_date",
    "end_date",
    "start_suspension_date",
    "next_invoice_date",
    "last_invoice_date",
    "months_to_next_invoice",
    "first_onb_date_ops",
    "onb_month_ops",
    "asset_activated_by",
    "asset_activated_by_email",
    "main_category",
    "initial_order_date",
    "subscription_management_new",
    "effective_pricing_date",
    "snap_start",
    "snap_end",
    "synced_at",
]


def write_to_bigquery(client: bigquery.Client, df: pd.DataFrame, dest_full: str) -> None:
    if df.empty:
        log.info("No rows to write to %s — skipping load.", dest_full)
        return

    log.info("Writing %d rows to %s ...", len(df), dest_full)

    out = df.copy()
    out["synced_at"] = pd.Timestamp.now(tz="UTC")

    existing = [c for c in ALL_INVOICES_COLUMNS if c in out.columns]
    out = out[existing]

    job_config = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_TRUNCATE,
        autodetect=True,
    )
    job = client.load_table_from_dataframe(out, dest_full, job_config=job_config)
    job.result()
    log.info("Write complete — %d rows in %s", len(out), dest_full)


# ── Bucket conditions — mirrors Node.js bucketCondition exactly ────────────────


def apply_bucket(df: pd.DataFrame, bucket: str, d: dict[str, str]) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    rd = pd.Timestamp(d["report_date"]).date()
    nd = pd.Timestamp(d["next_day"]).date()
    nnd = pd.Timestamp(d["next_next_day"]).date()
    nmrd = pd.Timestamp(d["next_month_report"]).date()

    lid = pd.to_datetime(df["last_invoice_date"], errors="coerce").dt.date
    nid = pd.to_datetime(df["next_invoice_date"], errors="coerce").dt.date

    late1 = (lid == nd) & (nid == nmrd)
    late2 = (lid == nnd) & (nid == nmrd)
    invoiced = (lid == rd) | late1 | late2

    inv = df["inv_revenue"].fillna(0)
    lpv = df["lpv_revenue"].fillna(0)
    ir = inv.round(2)
    lr = lpv.round(2)

    if bucket == "missing_inv":
        mask = (nid == rd) & (inv == 0) & (lpv > 0)
    elif bucket == "missing_lpv":
        mask = invoiced & (lpv == 0) & (inv > 0)
    elif bucket == "under":
        mask = invoiced & (ir < lr) & (lpv > 0) & (inv > 0)
    elif bucket == "over":
        mask = invoiced & (ir > lr) & (lpv > 0) & (inv > 0)
    else:
        raise ValueError(f"Unknown bucket: {bucket}")

    result = df[mask].copy()
    sort_cols = [c for c in ["sales_order", "cal_month"] if c in result.columns]
    if sort_cols:
        result = result.sort_values(sort_cols).reset_index(drop=True)
    return result


# ── Excel export — 4 sheets, matches emailAlert.js structure ──────────────────

SHEET_CONFIG = {
    "missing_inv": ("Missing in Invoice", "B45309"),
    "missing_lpv": ("Missing in OPS", "B45309"),
    "under": ("Under Invoice", "B45309"),
    "over": ("Over Invoice", "B45309"),
}

EXPORT_COLUMNS = [
    ("cal_month", "Billing Month"),
    ("country_code", "Country"),
    ("partner_id", "Partner ID"),
    ("establishment_uid", "Establishment"),
    ("sales_order", "Sales Order"),
    ("main_category", "Category"),
    ("product_code", "Product"),
    ("commitment_period", "Commitment Period"),
    ("billing_period", "Billing Period"),
    ("asset_activated_by", "Sales Agent"),
    ("lpv_revenue", "LPV (€)"),
    ("inv_revenue", "Invoice (€)"),
    ("abs_diff_eur", "Diff (€)"),
    ("discrepancy_risk", "Risk"),
    ("next_invoice_date", "Next Invoice"),
    ("end_date", "End Date"),
    ("start_date", "Start Date"),
    ("product_is_setup", "Invoice Type"),
    ("last_invoice_date", "Timestamp"),
    ("discrepancy_reason", "Discrepancy Reason"),
]

EUR_COLS = {"lpv_revenue", "inv_revenue", "abs_diff_eur"}
DATE_COLS = {"cal_month", "next_invoice_date", "end_date", "last_invoice_date", "start_date"}

FILL_WHITE = PatternFill("solid", fgColor="FFFFFFFF")
FILL_GRAY = PatternFill("solid", fgColor="FFF8FAFC")

HAIR = Side(style="hair")
HAIR_BORDER = Border(top=HAIR, bottom=HAIR, left=HAIR, right=HAIR)


def _write_sheet(ws: Any, df: pd.DataFrame, header_hex: str) -> None:
    hfill = PatternFill("solid", fgColor=header_hex)
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    dfont = Font(name="Arial", size=10)

    keys = [c[0] for c in EXPORT_COLUMNS]
    labels = [c[1] for c in EXPORT_COLUMNS]
    n_cols = len(labels)

    for ci, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill = hfill
        cell.font = hfont
        cell.border = HAIR_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"

    if df.empty:
        for ci, w in enumerate([14, 10, 18, 20, 18, 22, 18, 16, 22, 14, 12, 10, 14, 18], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        return

    for ri, row in enumerate(df.itertuples(index=False), 2):
        row_fill = FILL_WHITE if ri % 2 == 0 else FILL_GRAY

        for ci, key in enumerate(keys, 1):
            raw = getattr(row, key, None)

            if key == "product_is_setup":
                val = "Onetime" if raw is True else ("Recurring" if raw is False else "—")
            elif key in EUR_COLS:
                try:
                    val = float(raw) if raw is not None else 0.0
                except (TypeError, ValueError):
                    val = 0.0
            elif key in DATE_COLS:
                try:
                    val = pd.Timestamp(raw).date() if raw is not None else None
                except Exception:
                    val = None
            else:
                val = "" if (raw is None or (isinstance(raw, float) and pd.isna(raw))) else raw

            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = dfont
            cell.fill = row_fill
            cell.border = HAIR_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

            if key in EUR_COLS:
                cell.number_format = "#,##0.00 [$€-407]"

    last_row = len(df) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{last_row}"

    ref_widths = [14, 10, 18, 20, 18, 22, 18, 16, 22, 14, 12, 10, 14, 18]
    for ci, w in enumerate(ref_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.column_dimensions[get_column_letter(15)].width = 14
    ws.column_dimensions[get_column_letter(16)].width = 14
    ws.column_dimensions[get_column_letter(17)].width = 14
    ws.column_dimensions[get_column_letter(18)].width = 14
    ws.column_dimensions[get_column_letter(19)].width = 14
    ws.column_dimensions[get_column_letter(20)].width = 60


SUMMARY_COLUMNS = [
    "Type of issue",
    "Issue raised on",
    "Issue fixed on",
    "Country",
    "Product Family",
    "product_code",
    "recurring/onetime",
    "Count of invoices",
    "Est value (eur)",
    "Assignee",
    "Status",
    "Comments",
]

BUCKET_LABEL_MAP = {
    "missing_inv": "Missing in Invoice",
    "missing_lpv": "Missing in OPS",
    "under": "Under Invoice",
    "over": "Over Invoice",
}


def _write_summary_sheet(
    ws: Any, buckets: dict[str, pd.DataFrame], report_date_label: str
) -> None:
    """One row per country × product_code × invoice type, per discrepancy bucket."""
    hfill = PatternFill("solid", fgColor="B45309")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    dfont = Font(name="Arial", size=10)

    for ci, label in enumerate(SUMMARY_COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill = hfill
        cell.font = hfont
        cell.border = HAIR_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_COLUMNS))}1"

    rows: list[dict[str, Any]] = []
    for bucket, df in buckets.items():
        if df.empty:
            continue
        issue_label = BUCKET_LABEL_MAP.get(bucket, bucket)
        group_cols = ["country_code", "main_category", "product_code", "product_is_setup"]
        available = [c for c in group_cols if c in df.columns]
        if not available:
            continue
        grouped = df.groupby(available, dropna=False)
        for keys, grp in grouped:
            if not isinstance(keys, tuple):
                keys = (keys,)
            key_map = dict(zip(available, keys))
            is_setup = key_map.get("product_is_setup", None)
            if is_setup is True:
                inv_type = "Onetime"
            elif is_setup is False:
                inv_type = "Recurring"
            else:
                inv_type = "—"
            inv_col = grp["inv_revenue"].fillna(0)
            lpv_col = grp["lpv_revenue"].fillna(0)
            rows.append(
                {
                    "Type of issue": issue_label,
                    "Issue raised on": report_date_label,
                    "Issue fixed on": "",
                    "Country": key_map.get("country_code", ""),
                    "Product Family": key_map.get("main_category", ""),
                    "product_code": key_map.get("product_code", ""),
                    "recurring/onetime": inv_type,
                    "Count of invoices": len(grp),
                    "Est value (eur)": round(abs(lpv_col - inv_col).sum(), 2),
                    "Assignee": "",
                    "Status": "",
                    "Comments": "",
                }
            )

    rows.sort(key=lambda r: (r["Type of issue"], str(r["Country"]), str(r["product_code"])))

    for ri, row_data in enumerate(rows, 2):
        row_fill = FILL_WHITE if ri % 2 == 0 else FILL_GRAY
        for ci, col in enumerate(SUMMARY_COLUMNS, 1):
            val = row_data.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = dfont
            cell.fill = row_fill
            cell.border = HAIR_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
            if col == "Est value (eur)" and isinstance(val, (int, float)):
                cell.number_format = "#,##0.00 [$€-407]"

    last_row = max(len(rows) + 1, 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_COLUMNS))}{last_row}"
    widths = [20, 16, 16, 10, 16, 22, 14, 16, 16, 14, 14, 30]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def build_excel(buckets: dict[str, pd.DataFrame], report_date_label: str = "") -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    ws_summary = wb.create_sheet(title="Summary")
    _write_summary_sheet(ws_summary, buckets, report_date_label)
    for bucket, (sheet_name, color) in SHEET_CONFIG.items():
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, buckets[bucket], color)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Summary stats & template context ──────────────────────────────────────────


def _bucket_stats(df: pd.DataFrame) -> dict[str, Any]:
    inv = df["inv_revenue"].fillna(0) if not df.empty and "inv_revenue" in df.columns else pd.Series(dtype=float)
    lpv = df["lpv_revenue"].fillna(0) if not df.empty and "lpv_revenue" in df.columns else pd.Series(dtype=float)
    n = len(df)
    amt = abs(lpv - inv).sum() if n else 0.0
    return {
        "count": n,
        "amount": fmt_eur(amt),
        "avg": fmt_eur(amt / n if n else 0.0),
        "orders": df["sales_order"].nunique() if n else 0,
        "customers": df["establishment_uid"].nunique() if n else 0,
        "countries": df["country_code"].nunique() if n else 0,
    }


def build_template_context(
    df: pd.DataFrame,
    buckets: dict[str, pd.DataFrame],
    d: dict[str, str],
) -> dict[str, Any]:
    all_disc = (
        pd.concat(list(buckets.values()), ignore_index=True)
        if any(len(b) for b in buckets.values())
        else pd.DataFrame()
    )
    is_setup = all_disc.get("product_is_setup", pd.Series(dtype=object))

    mi = _bucket_stats(buckets["missing_inv"])
    ml = _bucket_stats(buckets["missing_lpv"])
    un = _bucket_stats(buckets["under"])
    ov = _bucket_stats(buckets["over"])

    return {
        "report_date_label": d["report_date_label"],
        "generated_date": datetime.now(BUSINESS_TZ).strftime("%d %B %Y"),
        "total_lpv_sales": fmt_eur(float(df["lpv_revenue"].fillna(0).sum()) if not df.empty else 0.0),
        "total_lpv_invoiced": fmt_eur(float(df["inv_revenue"].fillna(0).sum()) if not df.empty else 0.0),
        "missing_inv_count": mi["count"],
        "missing_inv_amount": mi["amount"],
        "missing_lpv_count": ml["count"],
        "missing_lpv_amount": ml["amount"],
        "under_count": un["count"],
        "under_amount": un["amount"],
        "over_count": ov["count"],
        "over_amount": ov["amount"],
        "recurring_count": int((is_setup == False).sum()),  # noqa: E712
        "onetime_count": int((is_setup == True).sum()),  # noqa: E712
    }


# ── Email helpers ─────────────────────────────────────────────────────────────


def render_html(context: dict[str, Any]) -> str:
    """Render HTML template using simple string replacement — no Jinja2 needed."""
    tmpl_path = TEMPLATE_DIR / TEMPLATE_FILE
    html = tmpl_path.read_text(encoding="utf-8")
    for key, value in context.items():
        html = html.replace("{{ " + key + " }}", str(value))
        html = html.replace("{{" + key + "}}", str(value))
    return html


def _parse_recipients(env_key: str, required: bool = True) -> list[str]:
    raw = _env(env_key, required=required)
    return [addr.strip() for addr in raw.split(",") if addr.strip()]


# ── Orchestration ─────────────────────────────────────────────────────────────


def generate_reports(output_dir: Path | str) -> list[dict[str, Any]]:
    """
    Run Invoice Radar, write Excel to ``output_dir``, optionally truncate-load
    ``bi.all_invoices``, and return a serializable email payload.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    tables = _tables()
    d = get_report_dates()
    log.info(
        "Invoice Radar starting — report date: %s  |  cal_month filter: %s",
        d["report_date"],
        d["cal_month_filter"],
    )

    client = get_bq_client(tables["project"])

    log.info("Fetching from lpv_adjusted ...")
    sql = build_source_sql(d, tables)
    df = run_query(client, sql, tables["project"])
    log.info("Fetched %d rows.", len(df))

    if "enriched_discrepancy_reason" in df.columns:
        df["discrepancy_reason"] = df["enriched_discrepancy_reason"].combine_first(
            df["discrepancy_reason"]
        )
        df.drop(columns=["enriched_discrepancy_reason"], inplace=True)
        log.info("Enriched discrepancy reasons applied.")

    if _write_all_invoices_enabled():
        write_to_bigquery(client, df, tables["dest_full"])
    else:
        log.info("WRITE_ALL_INVOICES is disabled — skipping BigQuery load.")

    buckets: dict[str, pd.DataFrame] = {
        b: apply_bucket(df, b, d) for b in ["missing_inv", "missing_lpv", "under", "over"]
    }
    for b, bdf in buckets.items():
        log.info("Bucket %-12s  %d rows", b, len(bdf))

    log.info("Building Excel ...")
    excel_bytes = build_excel(buckets, d["report_date_label"])
    safe_label = d["report_date_label"].replace(" ", "_")
    filename = f"invoice_radar_{safe_label}.xlsx"
    attachment_path = str(output_dir / filename)
    Path(attachment_path).write_bytes(excel_bytes)

    ctx = build_template_context(df, buckets, d)
    html_body = render_html(ctx)
    discrepancy_rows = sum(len(bdf) for bdf in buckets.values())

    payload = {
        "report_name": "Invoice Radar",
        "subject": f"Invoice Radar — Daily Report {d['report_date_label']}",
        "recipients": _parse_recipients("EMAIL_RECIPIENTS"),
        "cc": _parse_recipients("EMAIL_CC", required=False),
        "html_body": html_body,
        "attachment_path": attachment_path,
        "row_count": discrepancy_rows,
        "plain_text_summary": (
            f"Invoice Radar — {discrepancy_rows} discrepancy row(s) "
            f"for {d['report_date_label']}."
        ),
    }
    log.info("Generated Invoice Radar payload (%d discrepancy rows).", discrepancy_rows)
    return [payload]


def main() -> None:
    """Local entrypoint: generate reports then send via email_delivery."""
    try:
        from email_delivery import EmailDelivery, EmailMessage
    except ImportError:
        log.error(
            "Cannot import email_delivery. "
            "Run via Airflow or set PYTHONPATH to the dags folder."
        )
        raise

    staging = Path(os.getenv("INVOICE_RADAR_STAGING_DIR", "/tmp/invoice_radar"))
    payloads = generate_reports(staging)
    delivery = EmailDelivery.from_env()
    for payload in payloads:
        try:
            delivery.send(
                EmailMessage(
                    recipients=payload["recipients"],
                    subject=payload["subject"],
                    html_body=payload["html_body"],
                    cc=payload.get("cc") or None,
                    attachment_path=payload.get("attachment_path"),
                    plain_text_summary=payload.get("plain_text_summary"),
                )
            )
        except Exception:
            log.exception("Failed to send email for report '%s'.", payload["report_name"])

    log.info("Invoice Radar complete.")


if __name__ == "__main__":
    main()
